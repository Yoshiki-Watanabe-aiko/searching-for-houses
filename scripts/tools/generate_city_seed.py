"""総務省「全国地方公共団体コード」から ``db/seed/06_cities.sql`` を生成する。

``m_cities`` は YAML の ``search.cities`` に書く値の正典であり、市区必須サイトの
検索URLは JIS5桁コードから組み立てる。つまりこの表の欠落と誤りは
「その市区が丸ごと検索対象から漏れる」「別の市区のURLを叩く」に直結する。

正典を総務省のコード表へ移す理由（→ ADR 0014）:

* 初版はエイブルのエリア索引からの実測補完で、947件・15都道府県しか無かった
* 部分文字列一致で補完したため **他市のコードが混入していた**
  （名古屋市に北名古屋市の 23234、大阪市に東大阪市の 27227）
* 浜松市は2024年の区再編（7区→3区）に追随しておらず、新しい区名に
  旧区のコードが付いていた

使い方（いずれもネットワーク以外の副作用は無い）::

    # 1) 総務省のExcelを取得して中間CSVへ変換する（出典の更新時だけ）
    uv run python scripts/tools/generate_city_seed.py --fetch

    # 2) 中間CSVからシードSQLを生成する（既定）
    uv run python scripts/tools/generate_city_seed.py

    # 3) 生成せず、DBの現状との差分だけを報告する
    uv run python scripts/tools/generate_city_seed.py --report

⚠ **中間CSVを正典として Git 管理する。** Excel はバイナリで差分が読めず、
総務省の配布URLは改訂のたびに変わるため、取得済みのCSVを版として残さないと
「いつ時点のコード表で生成したか」を後から言えなくなる。
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
CSV_PATH = PROJECT_ROOT / "data" / "city_master" / "soumu_local_gov_codes.csv"
SQL_PATH = PROJECT_ROOT / "db" / "seed" / "06_cities.sql"

# 総務省「全国地方公共団体コード」の配布ページと、そこに載っている団体コード表。
# ⚠ 改訂のたびにファイル名（数字）が変わる。--fetch はページを読んで
# 「団体コード表」のリンクを自力で見つけるのではなく、ここを更新して使う。
SOURCE_PAGE = "https://www.soumu.go.jp/denshijiti/code.html"
SOURCE_XLSX = "https://www.soumu.go.jp/main_content/000925835.xlsx"
SHEET_MUNICIPALITIES = "R6.1.1現在の団体"
SHEET_DESIGNATED = "R6.1.1政令指定都市"
SOURCE_AS_OF = "2024-01-01"  # コード表の基準日（シート名の R6.1.1）

# 同一都道府県内で ``canonical_name`` が衝突する唯一の組。
# 北海道には後志総合振興局 古宇郡泊村（01403）と根室振興局 国後郡泊村（01696）が
# あり、市区町村名だけでは区別できない。政令市の区を市名で prefix するのと
# 同じ考え方で、衝突する側だけ郡名を前置して一意にする。
CANONICAL_OVERRIDES = {"01696": "国後郡泊村"}


@dataclass(frozen=True, slots=True)
class CityRow:
    """シード1行ぶん。``m_cities`` の列にそのまま対応する。"""

    jis_code: str
    prefecture: str
    city_name: str
    parent_city: str | None
    canonical_name: str


def normalize_name(name: str) -> str:
    """突き合わせ用に市区町村名を正規化する。

    NFKC に加えて、自治体名で実際に揺れる異体字を寄せる。マスタ側の表記を
    書き換えるためではなく、**照合を落とさない**ためだけに使う。
    """
    return unicodedata.normalize("NFKC", name).translate(
        str.maketrans("ヶヵ﨑德瀧曽舘", "ケカ崎徳滝曾館")
    )


def fetch_from_soumu() -> list[CityRow]:
    """総務省のExcelを取得して行に変換する（``--fetch`` のときだけ呼ぶ）。"""
    import io

    import httpx
    import openpyxl

    response = httpx.get(SOURCE_XLSX, timeout=60.0, follow_redirects=True)
    response.raise_for_status()
    workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)

    # シート1は都道府県と市区町村。市区町村名が空の行は都道府県そのものなので落とす。
    municipalities: list[tuple[str, str, str]] = []
    for code, prefecture, city, *_ in workbook[SHEET_MUNICIPALITIES].iter_rows(
        min_row=2, values_only=True
    ):
        if code and city:
            municipalities.append((str(code)[:5], prefecture.strip(), city.strip()))

    # シート2は政令指定都市。市そのものの行と行政区の行が混在している。
    designated: list[tuple[str, str, str]] = []
    for code, prefecture, city, *_ in workbook[SHEET_DESIGNATED].iter_rows(
        min_row=2, values_only=True
    ):
        if code and city:
            designated.append((str(code)[:5], prefecture.strip(), city.strip()))
    workbook.close()

    return _build_rows(municipalities, designated)


def _build_rows(
    municipalities: list[tuple[str, str, str]], designated: list[tuple[str, str, str]]
) -> list[CityRow]:
    """総務省の2シートを ``m_cities`` の表現へ組み替える。

    行政区は「札幌市中央区」のようにフルネームで載っているので、市名で分割して
    ``parent_city`` と ``city_name`` に割る（既存マスタの表現に合わせる）。
    """
    known = {(prefecture, city) for _, prefecture, city in municipalities}
    rows = [
        CityRow(code, prefecture, city, None, city)
        for code, prefecture, city in municipalities
    ]

    # 長い市名から順に当てる。「川崎市川崎区」を「川崎市」で切るため。
    parents = sorted(
        {city for _, _, city in municipalities if city.endswith("市")},
        key=len,
        reverse=True,
    )
    for code, prefecture, full_name in designated:
        if (prefecture, full_name) in known:
            continue  # 市そのものの行はシート1に既にある
        parent = next(
            (p for p in parents if full_name.startswith(p) and full_name != p), None
        )
        if parent is None:
            raise ValueError(f"行政区の市名を特定できません: {prefecture} {full_name}")
        rows.append(
            CityRow(code, prefecture, full_name[len(parent) :], parent, full_name)
        )

    for index, row in enumerate(rows):
        override = CANONICAL_OVERRIDES.get(row.jis_code)
        if override:
            rows[index] = CityRow(
                row.jis_code, row.prefecture, row.city_name, row.parent_city, override
            )

    rows.sort(key=lambda r: r.jis_code)
    _assert_unique(rows)
    return rows


def _assert_unique(rows: list[CityRow]) -> None:
    """シードを流す前に、DBの一意制約を破る組が無いことを確かめる。

    ``m_cities`` は (prefecture, canonical_name) が UNIQUE で、
    ADR 0014 で jis_code にも部分ユニーク索引を張った。生成時に落とさないと
    ``db-seed`` の途中で失敗して原因が分かりにくくなる。
    """
    by_name: dict[tuple[str, str], list[str]] = {}
    by_code: dict[str, int] = {}
    for row in rows:
        by_name.setdefault((row.prefecture, row.canonical_name), []).append(row.jis_code)
        by_code[row.jis_code] = by_code.get(row.jis_code, 0) + 1
    name_dups = {k: v for k, v in by_name.items() if len(v) > 1}
    code_dups = {k: v for k, v in by_code.items() if v > 1}
    if name_dups:
        raise ValueError(f"(都道府県, canonical_name) が重複しています: {name_dups}")
    if code_dups:
        raise ValueError(f"jis_code が重複しています: {code_dups}")


def write_csv(rows: list[CityRow], path: Path = CSV_PATH) -> None:
    """中間CSV（Git管理下の正典）を書き出す。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(["jis_code", "prefecture", "city_name", "parent_city", "canonical_name"])
        for row in rows:
            writer.writerow(
                [
                    row.jis_code,
                    row.prefecture,
                    row.city_name,
                    row.parent_city or "",
                    row.canonical_name,
                ]
            )


def read_csv(path: Path = CSV_PATH) -> list[CityRow]:
    """中間CSVを読む。"""
    with path.open(encoding="utf-8", newline="") as handle:
        return [
            CityRow(
                r["jis_code"],
                r["prefecture"],
                r["city_name"],
                r["parent_city"] or None,
                r["canonical_name"],
            )
            for r in csv.DictReader(handle)
        ]


def _sql_literal(value: str | None) -> str:
    """SQLの文字列リテラルにする。自治体名にアポストロフィは出ないが一応エスケープする。"""
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def render_sql(rows: list[CityRow]) -> str:
    """冪等なシードSQLを組み立てる。

    ``ON CONFLICT (prefecture, canonical_name) DO UPDATE`` で、既存行の
    ``jis_code`` を**総務省の値で上書きする**。COALESCE で温存すると
    混入した誤コード（名古屋市の 23234 等）が永久に残るため、ここは
    「総務省が正」と決め切る。
    """
    prefectures = len({row.prefecture for row in rows})
    wards = sum(1 for row in rows if row.parent_city)
    header = f"""\
-- ============================================================
-- m_cities: 市区町村マスタ（全国版）
--
-- **このファイルは生成物**。手で編集せず、
--   uv run python scripts/tools/generate_city_seed.py
-- で作り直す。正典は data/city_master/soumu_local_gov_codes.csv。
--
-- 出典: 総務省「全国地方公共団体コード」
--   {SOURCE_PAGE}
--   基準日 {SOURCE_AS_OF}（コード表シート「{SHEET_MUNICIPALITIES}」）
--   政府標準利用規約に基づき出典を明示して利用する。
--
-- 行数: {len(rows)}行 = 市区町村{len(rows) - wards}（東京23区と北方領土6村を含む）
--       ＋ 政令指定都市の行政区{wards}。{prefectures}都道府県。
--
-- canonical_name は YAML の search.cities に書く値の正典で、同一都道府県内で
-- 一意。政令市の区は市名を prefix する（例: 横浜市西区）。
-- ⚠ 北海道の泊村だけは市区町村どうしで衝突する（後志 01403 / 根室 01696）ため、
--   北方領土側を「国後郡泊村」として区別している。
--
-- 【この全国版へ移行した理由 → ADR 0014】
-- 初版はエイブルのエリア索引からの実測補完で 947行・15都道府県しか無く、
-- 市区必須サイトでは登録の無い市区が丸ごと検索対象から漏れていた（課題#16）。
-- さらに部分文字列一致で補完したため、実測で次の5件の誤りが判明した。
-- いずれも本ファイルの適用で総務省の値へ訂正される。
--   静岡県 浜松市中央区  22131 → 22138  ）2024年の区再編（7区→3区）に
--   静岡県 浜松市浜名区  22132 → 22139  ）追随しておらず、新しい区名に
--   静岡県 浜松市天竜区  22133 → 22140  ）旧区のコードが付いていた
--   愛知県 名古屋市      23234 → 23100   北名古屋市のコードが混入していた
--   大阪府 大阪市        27227 → 27100   東大阪市のコードが混入していた
--
-- ⚠ 廃置分合で消えた自治体の行は**削除しない**。t_listings.city_id から
--   参照されている可能性があるため、総務省一覧に無い既存行は残す。
-- ============================================================

INSERT INTO m_cities (prefecture, parent_city, city_name, canonical_name, jis_code) VALUES
"""
    values = ",\n".join(
        f"    ({_sql_literal(row.prefecture)}, {_sql_literal(row.parent_city)},"
        f" {_sql_literal(row.city_name)}, {_sql_literal(row.canonical_name)},"
        f" {_sql_literal(row.jis_code)})"
        for row in rows
    )
    footer = """
ON CONFLICT (prefecture, canonical_name) DO UPDATE SET
    parent_city = EXCLUDED.parent_city,
    city_name   = EXCLUDED.city_name,
    -- COALESCE で温存しない。混入した誤コードを上書きするのが移行の目的。
    jis_code    = EXCLUDED.jis_code,
    updated_at  = now();
"""
    return header + values + footer


def report_against_db(rows: list[CityRow]) -> int:
    """DBの現状と生成データの差分を報告する。戻り値は終了コード。"""
    from sqlalchemy import text

    from house_search.db.session import get_engine

    by_key = {(row.prefecture, normalize_name(row.canonical_name)): row for row in rows}
    with get_engine().connect() as conn:
        existing = list(
            conn.execute(
                text(
                    "SELECT prefecture, parent_city, city_name, canonical_name, jis_code"
                    " FROM m_cities"
                )
            )
        )

    unmatched = [
        r for r in existing if (r.prefecture, normalize_name(r.canonical_name)) not in by_key
    ]
    # NULL の補完と、値の入れ替わり（＝誤コードの訂正）は分けて数える。
    # 混ぜると数百件の補完に紛れて、訂正すべき数件が見えなくなる。
    filled: list[tuple[str, str, str]] = []
    corrected: list[tuple[str, str, str, str]] = []
    for row in existing:
        key = (row.prefecture, normalize_name(row.canonical_name))
        if key not in by_key:
            continue
        want = by_key[key].jis_code
        if row.jis_code is None:
            filled.append((row.prefecture, row.canonical_name, want))
        elif row.jis_code != want:
            corrected.append((row.prefecture, row.canonical_name, row.jis_code, want))

    existing_keys = {(r.prefecture, normalize_name(r.canonical_name)) for r in existing}
    added = [k for k in by_key if k not in existing_keys]

    print(f"DB既存 {len(existing)}行 / 生成 {len(rows)}行")
    print(f"  新規に追加される行        : {len(added)}")
    print(f"  jis_code が NULL から埋まる: {len(filled)}")
    print(f"  jis_code が訂正される行    : {len(corrected)}")
    for fix in corrected:
        print(f"    {fix[0]} {fix[1]}: {fix[2]} -> {fix[3]}")
    print(f"  総務省一覧に無い既存行    : {len(unmatched)}（削除せず残す）")
    for row in unmatched:
        print(f"    {row.prefecture} {row.canonical_name} ({row.jis_code})")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--fetch",
        action="store_true",
        help="総務省のExcelを取得して中間CSVを作り直す（出典の更新時だけ）",
    )
    parser.add_argument(
        "--report",
        action="store_true",
        help="SQLを書かず、DBの現状との差分だけを報告する",
    )
    args = parser.parse_args(argv)

    if args.fetch:
        rows = fetch_from_soumu()
        write_csv(rows)
        print(f"中間CSVを更新しました: {CSV_PATH}（{len(rows)}行・取得日 {dt.date.today()}）")
    else:
        rows = read_csv()
        _assert_unique(rows)

    if args.report:
        return report_against_db(rows)

    SQL_PATH.write_text(render_sql(rows), encoding="utf-8", newline="\n")
    print(f"シードSQLを生成しました: {SQL_PATH}（{len(rows)}行）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
